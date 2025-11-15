from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from proyectos.models import Proyecto, ParticipacionProyecto
from proyectos.forms import ProyectoCrearForm
from roles.models import Rol
from accounts.models import Usuario
from grupos.models import Grupo
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse

# Importar vista de exportación PDF del dashboard
from proyectos.views_dashboard_pdf import exportar_dashboard_pdf


# Helpers
def is_admin(user):
    return user.roles.filter(nombre__iexact="Admin").exists()


def actualizar_rol_lider(usuario):
    """
    Actualiza automáticamente el rol de un usuario según si lidera proyectos o no.
    
    Reglas:
    - Si el usuario lidera AL MENOS UN proyecto → rol "Líder"
    - Si el usuario NO lidera ningún proyecto → rol "Desarrollador"
    - Los usuarios con rol "Admin" NO se ven afectados (no pueden ser líderes)
    """
    # No modificar el rol de administradores
    if usuario.roles.filter(nombre__iexact="Admin").exists():
        return
    
    # Obtener roles
    rol_lider = Rol.objects.get(nombre="Líder")
    rol_developer = Rol.objects.get(nombre="Desarrollador")
    
    # Verificar si el usuario lidera algún proyecto
    lidera_proyectos = Proyecto.objects.filter(lider=usuario).exists()
    
    if lidera_proyectos:
        # Debe tener rol "Líder"
        if not usuario.roles.filter(nombre="Líder").exists():
            # Cambiar de Desarrollador → Líder
            usuario.roles.remove(rol_developer)
            usuario.roles.add(rol_lider)
    else:
        # NO lidera proyectos, debe tener rol "Desarrollador"
        if usuario.roles.filter(nombre="Líder").exists():
            # Cambiar de Líder → Desarrollador
            usuario.roles.remove(rol_lider)
            usuario.roles.add(rol_developer)


def usuario_tiene_proyecto_activo(usuario, proyecto_actual=None):
    """
    Verifica si un usuario ya está asignado a un proyecto activo.
    
    Args:
        usuario: El usuario a verificar
        proyecto_actual: Proyecto a excluir de la verificación (para edición)
    
    Returns:
        tuple: (tiene_proyecto_activo, nombre_proyecto)
    """
    # Buscar participaciones del usuario en proyectos activos
    participaciones = ParticipacionProyecto.objects.filter(
        usuario=usuario,
        proyecto__activo=True
    ).select_related('proyecto')
    
    # Excluir el proyecto actual si se está editando
    if proyecto_actual:
        participaciones = participaciones.exclude(proyecto=proyecto_actual)
    
    if participaciones.exists():
        proyecto_activo = participaciones.first().proyecto
        return True, proyecto_activo.nombre
    
    return False, None

@login_required
@user_passes_test(is_admin)
def lista_proyectos(request):
    proyectos_qs = Proyecto.objects.select_related('lider', 'grupo', 'creado_por').prefetch_related('clientes')
    sort = request.GET.get('sort', '')
    
    # Orden por defecto: último creado primero (por ID descendente)
    if not sort:
        proyectos_qs = proyectos_qs.order_by('-id')
    elif sort == 'nombre':
        proyectos_qs = proyectos_qs.order_by('nombre')
    elif sort == 'lider':
        proyectos_qs = proyectos_qs.order_by('lider__nombre')
    elif sort == 'creado_por':
        proyectos_qs = proyectos_qs.order_by('creado_por__nombre')
    elif sort == 'fecha_creacion':
        proyectos_qs = proyectos_qs.order_by('-fecha_creacion')
    
    # Paginación
    paginator = Paginator(proyectos_qs, 10)
    page_number = request.GET.get('page')
    proyectos = paginator.get_page(page_number)
    
    return render(request, "proyectos/lista_proyectos.html", {
        "proyectos": proyectos,
        "sort": sort,
        "page_title": "Lista de Proyectos"
    })


@login_required
@user_passes_test(is_admin)
def buscar_proyectos_ajax(request):
    """Endpoint AJAX para búsqueda de proyectos"""
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse({'proyectos': [], 'count': 0})

    # Buscar por nombre, líder o clientes
    proyectos = Proyecto.objects.filter(
        Q(nombre__icontains=q) |
        Q(lider__nombre__icontains=q) |
        Q(clientes__nombre__icontains=q)
    ).select_related('lider', 'creado_por').prefetch_related('clientes').distinct()[:50]

    proyectos_data = []
    for p in proyectos:
        proyectos_data.append({
            'id': p.id,
            'nombre': p.nombre,
            'logo': p.logo.url if p.logo else None,
            'lider': p.lider.nombre if p.lider else None,
            'clientes': [c.nombre for c in p.clientes.all()],
            'creado_por': p.creado_por.nombre if p.creado_por else None,
            'fecha_creacion': p.fecha_creacion.strftime('%d/%m/%Y %H:%M') if hasattr(p, 'fecha_creacion') and p.fecha_creacion else ''
        })

    return JsonResponse({'proyectos': proyectos_data, 'count': len(proyectos_data)})


@login_required
@user_passes_test(is_admin)
def crear_proyecto(request):
    if request.method == "POST":
        form = ProyectoCrearForm(request.POST, request.FILES)
        if form.is_valid():
            # Crear el proyecto
            proyecto = form.save(commit=False)
            proyecto.creado_por = request.user
            proyecto.save()

            # Obtener el grupo seleccionado (puede ser None)
            grupo = proyecto.grupo

            if grupo:
                # ⚡ VALIDAR: Un grupo solo puede tener UN proyecto activo
                proyecto_existente = Proyecto.objects.filter(grupo=grupo, activo=True).first()
                if proyecto_existente:
                    proyecto.delete()  # Eliminar el proyecto creado
                    messages.error(
                        request,
                        f"El grupo '{grupo.nombre}' ya tiene asignado el proyecto activo '{proyecto_existente.nombre}'. "
                        f"Un grupo solo puede tener un proyecto activo a la vez."
                    )
                    return render(request, "proyectos/crear_proyecto.html", {
                        "form": form,
                        "page_title": "Crear Proyecto"
                    })
                
                # Solo procesar líder y participantes si hay grupo
                lider_id = form.cleaned_data.get('lider')
                if lider_id:
                    lider = Usuario.objects.get(id=lider_id)

                    # Validar que el líder no sea Admin
                    if lider.roles.filter(nombre__iexact="Admin").exists():
                        proyecto.delete()  # Eliminar el proyecto creado
                        messages.error(request, "Un usuario con rol 'Admin' no puede ser líder de proyecto.")
                        return render(request, "proyectos/crear_proyecto.html", {
                            "form": form,
                            "page_title": "Crear Proyecto"
                        })
                    
                    # ⚡ VALIDAR: El líder no puede estar en otro proyecto activo
                    tiene_proyecto, nombre_proyecto = usuario_tiene_proyecto_activo(lider)
                    if tiene_proyecto:
                        proyecto.delete()  # Eliminar el proyecto creado
                        messages.error(
                            request,
                            f"El usuario '{lider.nombre}' ya está asignado al proyecto activo '{nombre_proyecto}'. "
                            f"Una persona solo puede estar en un proyecto activo a la vez."
                        )
                        return render(request, "proyectos/crear_proyecto.html", {
                            "form": form,
                            "page_title": "Crear Proyecto"
                        })

                    # Crear rol "Líder" si no existe
                    rol_lider, _ = Rol.objects.get_or_create(
                        nombre="Líder",
                        defaults={"color": "#28a745"}
                    )

                    # Asignar líder al proyecto
                    proyecto.lider = lider
                    proyecto.save()

                    # ⚡ ACTUALIZAR ROL DEL LÍDER (Developer → Líder)
                    actualizar_rol_lider(lider)

                    # Agregar líder con rol "Líder"
                    ParticipacionProyecto.objects.create(
                        usuario=lider,
                        proyecto=proyecto,
                        rol=rol_lider
                    )

                    # Agregar todos los demás integrantes del grupo como "Desarrollador"
                    rol_dev, _ = Rol.objects.get_or_create(
                        nombre="Desarrollador",
                        defaults={"color": "#ffc107"}
                    )

                    # ⚡ VALIDAR: Cada integrante no puede estar en otro proyecto activo
                    integrantes_con_conflicto = []
                    for integrante in grupo.integrantes.exclude(id=lider_id):
                        tiene_proyecto, nombre_proyecto = usuario_tiene_proyecto_activo(integrante)
                        if tiene_proyecto:
                            integrantes_con_conflicto.append(f"{integrante.nombre} (en '{nombre_proyecto}')")
                        else:
                            ParticipacionProyecto.objects.create(
                                usuario=integrante,
                                proyecto=proyecto,
                                rol=rol_dev
                            )
                    
                    if integrantes_con_conflicto:
                        messages.warning(
                            request,
                            f"Los siguientes integrantes ya están en proyectos activos y no fueron agregados: "
                            f"{', '.join(integrantes_con_conflicto)}"
                        )

                messages.success(request, f"Proyecto '{proyecto.nombre}' creado exitosamente con el grupo '{grupo.nombre}'.")
            else:
                # Proyecto sin grupo
                messages.success(request, f"Proyecto '{proyecto.nombre}' creado exitosamente sin grupo asignado.")

            # Asignar clientes al proyecto (independientemente de si hay grupo o no)
            clientes_ids = form.cleaned_data.get('clientes')
            if clientes_ids:
                # Convertir IDs (strings) a enteros y obtener objetos Usuario
                clientes_ids_int = [int(id) for id in clientes_ids]
                clientes = Usuario.objects.filter(id__in=clientes_ids_int)
                proyecto.clientes.set(clientes)
                # Asignar rol Stakeholder en ParticipacionProyecto para cada cliente
                rol_stakeholder, _ = Rol.objects.get_or_create(
                    nombre="Stakeholder",
                    defaults={"color": "#17a2b8"}
                )
                for cliente in clientes:
                    ParticipacionProyecto.objects.get_or_create(
                        usuario=cliente,
                        proyecto=proyecto,
                        defaults={'rol': rol_stakeholder}
                    )

            return redirect("proyectos:lista_proyectos")
    else:
        form = ProyectoCrearForm()

    return render(request, "proyectos/crear_proyecto.html", {
        "form": form,
        "page_title": "Crear Proyecto"
    })


@login_required
@user_passes_test(is_admin)
def editar_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == "POST":
        form = ProyectoCrearForm(request.POST, request.FILES, instance=proyecto)
        if form.is_valid():
            proyecto = form.save()

            # Obtener el grupo seleccionado (puede ser None)
            grupo = proyecto.grupo

            if grupo:
                # ⚡ VALIDAR: Un grupo solo puede tener UN proyecto activo (excepto el proyecto actual)
                proyecto_existente = Proyecto.objects.filter(grupo=grupo, activo=True).exclude(id=proyecto.id).first()
                if proyecto_existente:
                    messages.error(
                        request,
                        f"El grupo '{grupo.nombre}' ya tiene asignado el proyecto activo '{proyecto_existente.nombre}'. "
                        f"Un grupo solo puede tener un proyecto activo a la vez."
                    )
                    return render(request, "proyectos/editar_proyecto.html", {
                        "form": form,
                        "proyecto": proyecto,
                        "page_title": "Editar Proyecto"
                    })
                
                # Solo procesar líder y participantes si hay grupo
                lider_id = form.cleaned_data.get('lider')
                if lider_id:
                    lider = Usuario.objects.get(id=lider_id)
                    
                    # Validar que el líder no sea Admin
                    if lider.roles.filter(nombre__iexact="Admin").exists():
                        messages.error(request, "Un usuario con rol 'Admin' no puede ser líder de proyecto.")
                        return render(request, "proyectos/editar_proyecto.html", {
                            "form": form,
                            "proyecto": proyecto,
                            "page_title": "Editar Proyecto"
                        })
                    
                    # ⚡ VALIDAR: El líder no puede estar en otro proyecto activo (excepto este)
                    tiene_proyecto, nombre_proyecto = usuario_tiene_proyecto_activo(lider, proyecto)
                    if tiene_proyecto:
                        messages.error(
                            request,
                            f"El usuario '{lider.nombre}' ya está asignado al proyecto activo '{nombre_proyecto}'. "
                            f"Una persona solo puede estar en un proyecto activo a la vez."
                        )
                        return render(request, "proyectos/editar_proyecto.html", {
                            "form": form,
                            "proyecto": proyecto,
                            "page_title": "Editar Proyecto"
                        })

                    # Guardar el líder anterior para actualizar su rol después
                    lider_anterior = proyecto.lider

                    # Crear rol "Líder" si no existe
                    rol_lider, _ = Rol.objects.get_or_create(
                        nombre="Líder",
                        defaults={"color": "#28a745"}
                    )

                    # Asignar líder al proyecto
                    proyecto.lider = lider
                    proyecto.save()

                    # ⚡ ACTUALIZAR ROL DEL NUEVO LÍDER (Developer → Líder)
                    actualizar_rol_lider(lider)

                    # ⚡ ACTUALIZAR ROL DEL LÍDER ANTERIOR (Líder → Developer si ya no lidera proyectos)
                    if lider_anterior and lider_anterior != lider:
                        actualizar_rol_lider(lider_anterior)

                    # Limpiar participantes actuales
                    proyecto.participantes.clear()

                    # Agregar líder con rol "Líder"
                    ParticipacionProyecto.objects.create(
                        usuario=lider,
                        proyecto=proyecto,
                        rol=rol_lider
                    )

                    # Agregar todos los demás integrantes del grupo como "Desarrollador"
                    rol_dev, _ = Rol.objects.get_or_create(
                        nombre="Desarrollador",
                        defaults={"color": "#ffc107"}
                    )

                    # ⚡ VALIDAR: Cada integrante no puede estar en otro proyecto activo
                    integrantes_con_conflicto = []
                    for integrante in grupo.integrantes.exclude(id=lider_id):
                        tiene_proyecto, nombre_proyecto = usuario_tiene_proyecto_activo(integrante, proyecto)
                        if tiene_proyecto:
                            integrantes_con_conflicto.append(f"{integrante.nombre} (en '{nombre_proyecto}')")
                        else:
                            ParticipacionProyecto.objects.create(
                                usuario=integrante,
                                proyecto=proyecto,
                                rol=rol_dev
                            )
                    
                    if integrantes_con_conflicto:
                        messages.warning(
                            request,
                            f"Los siguientes integrantes ya están en proyectos activos y no fueron agregados: "
                            f"{', '.join(integrantes_con_conflicto)}"
                        )

                messages.success(request, f"Proyecto '{proyecto.nombre}' actualizado exitosamente con el grupo '{grupo.nombre}'.")
            else:
                # Proyecto sin grupo - guardar líder anterior y limpiar
                lider_anterior = proyecto.lider
                
                proyecto.lider = None
                proyecto.save()
                proyecto.participantes.clear()
                
                # ⚡ ACTUALIZAR ROL DEL LÍDER ANTERIOR (Líder → Developer si ya no lidera proyectos)
                if lider_anterior:
                    actualizar_rol_lider(lider_anterior)
                
                messages.success(request, f"Proyecto '{proyecto.nombre}' actualizado exitosamente sin grupo asignado.")

            # Actualizar clientes del proyecto (independientemente de si hay grupo o no)
            clientes_ids = form.cleaned_data.get('clientes')
            if clientes_ids:
                # Convertir IDs (strings) a enteros y obtener objetos Usuario
                clientes_ids_int = [int(id) for id in clientes_ids]
                clientes = Usuario.objects.filter(id__in=clientes_ids_int)
                
                # Obtener clientes anteriores para limpiar ParticipacionProyecto
                clientes_anteriores = set(proyecto.clientes.all())
                clientes_nuevos = set(clientes)
                
                # Eliminar ParticipacionProyecto de clientes que ya no están asignados
                clientes_removidos = clientes_anteriores - clientes_nuevos
                for cliente in clientes_removidos:
                    ParticipacionProyecto.objects.filter(
                        usuario=cliente,
                        proyecto=proyecto,
                        rol__nombre='Stakeholder'
                    ).delete()
                
                # Actualizar la lista de clientes
                proyecto.clientes.set(clientes)
                
                # Asignar rol Stakeholder en ParticipacionProyecto para nuevos clientes
                rol_stakeholder, _ = Rol.objects.get_or_create(
                    nombre="Stakeholder",
                    defaults={"color": "#17a2b8"}
                )
                for cliente in clientes:
                    ParticipacionProyecto.objects.get_or_create(
                        usuario=cliente,
                        proyecto=proyecto,
                        defaults={'rol': rol_stakeholder}
                    )
            else:
                # Si no hay clientes seleccionados, limpiar todos
                clientes_anteriores = proyecto.clientes.all()
                for cliente in clientes_anteriores:
                    ParticipacionProyecto.objects.filter(
                        usuario=cliente,
                        proyecto=proyecto,
                        rol__nombre='Stakeholder'
                    ).delete()
                proyecto.clientes.clear()

            return redirect("proyectos:lista_proyectos")
    else:
        form = ProyectoCrearForm(instance=proyecto)

    return render(request, "proyectos/editar_proyecto.html", {
        "form": form,
        "proyecto": proyecto,
        "page_title": "Editar Proyecto"
    })


@login_required
@user_passes_test(is_admin)
def eliminar_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == "POST":
        # Guardar el líder antes de eliminar el proyecto
        lider_anterior = proyecto.lider
        
        proyecto.delete()
        
        # ⚡ ACTUALIZAR ROL DEL LÍDER (Líder → Developer si ya no lidera proyectos)
        if lider_anterior:
            actualizar_rol_lider(lider_anterior)
        
        messages.success(request, f"Proyecto '{proyecto.nombre}' eliminado correctamente.")
        return redirect("proyectos:lista_proyectos")

    return render(request, "proyectos/confirmar_eliminar_proyecto.html", {
        "proyecto": proyecto,
        "page_title": "Eliminar Proyecto"
    })


def is_lider_del_proyecto(user, proyecto):
    """Helper para verificar si el usuario es líder del proyecto"""
    return proyecto.lider == user


@login_required
def asignar_metodologia(request, proyecto_id):
    """
    Vista para que el líder asigne o cambie la metodología del proyecto.
    Muestra advertencia si hay requerimientos o casos de uso, pero permite ver el formulario.
    """
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Verificar que el usuario sea el líder del proyecto
    if not is_lider_del_proyecto(request.user, proyecto):
        messages.error(request, "No tienes permiso para modificar este proyecto. Solo el líder puede asignar la metodología.")
        return redirect("dashboards:lider_dashboard")
    
    puede_cambiar = proyecto.puede_cambiar_metodologia()
    necesita_metodologia = proyecto.necesita_metodologia()
    
    if request.method == "POST":
        # Validar nuevamente en el POST si se puede cambiar
        if not necesita_metodologia and not puede_cambiar:
            messages.error(
                request, 
                "No puedes cambiar la metodología porque el proyecto ya tiene requerimientos o casos de uso cargados. "
                "Elimina primero todos los requerimientos y casos de uso para poder cambiar la metodología."
            )
            return redirect("proyectos:asignar_metodologia", proyecto_id=proyecto_id)
        
        metodologia = request.POST.get('metodologia')
        
        if metodologia in ['TRADICIONAL', 'AGIL']:
            proyecto.metodologia = metodologia
            proyecto.save()
            
            metodologia_nombre = "Tradicional" if metodologia == "TRADICIONAL" else "Ágil"
            messages.success(request, f"Metodología '{metodologia_nombre}' asignada correctamente al proyecto '{proyecto.nombre}'.")
            return redirect("dashboards:lider_dashboard")
        else:
            messages.error(request, "Metodología inválida seleccionada.")
    
    return render(request, "proyectos/asignar_metodologia.html", {
        "proyecto": proyecto,
        "puede_cambiar": puede_cambiar,
        "necesita_metodologia": necesita_metodologia,
        "page_title": "Asignar Metodología"
    })


@login_required
def matriz_trazabilidad(request, proyecto_id):
    """
    Vista de Matriz de Trazabilidad con Live Traceability.
    Muestra relaciones dinámicas entre requerimientos y casos de uso.
    Accesible por líder y participantes del proyecto.
    """
    from requerimientos.models import Requerimiento, RequerimientoCaso
    from casos_de_uso.models import CasoDeUso
    from django.db.models import Prefetch, Count
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Verificar permisos: líder o participante del proyecto
    es_lider = proyecto.lider == request.user
    es_participante = proyecto.participantes.filter(id=request.user.id).exists()
    
    if not (es_lider or es_participante):
        messages.error(request, "No tienes permiso para ver la matriz de trazabilidad de este proyecto.")
        return redirect("dashboards:lider_dashboard")
    
    # Filtros
    tipo_req = request.GET.get('tipo_req')
    estado_req = request.GET.get('estado_req')
    solo_huerfanos = request.GET.get('solo_huerfanos') == 'true'
    solo_sin_cubrir = request.GET.get('solo_sin_cubrir') == 'true'
    
    # Obtener requerimientos con filtros (excluir NO FUNCIONALES y BORRADOR)
    # Solo se muestran requerimientos VALIDADOS o en estados posteriores
    requerimientos_qs = Requerimiento.objects.filter(
        proyecto=proyecto, 
        tipo='FUNCIONAL'
    ).exclude(estado='BORRADOR')
    
    if tipo_req:
        requerimientos_qs = requerimientos_qs.filter(tipo=tipo_req)
    if estado_req:
        requerimientos_qs = requerimientos_qs.filter(estado=estado_req)
    if solo_huerfanos:
        requerimientos_qs = requerimientos_qs.annotate(
            rel_count=Count('relaciones_casos')
        ).filter(rel_count=0)
    
    # Prefetch optimizado para relaciones
    requerimientos = requerimientos_qs.prefetch_related(
        Prefetch('relaciones_casos',
                queryset=RequerimientoCaso.objects.select_related('caso_de_uso'))
    ).order_by('id')
    
    # Obtener casos de uso del proyecto
    casos = CasoDeUso.objects.filter(proyecto=proyecto).prefetch_related(
        Prefetch('relaciones_requerimientos',
                queryset=RequerimientoCaso.objects.select_related('requerimiento'))
    ).order_by('id')
    
    if solo_sin_cubrir:
        casos = casos.annotate(rel_count=Count('relaciones_requerimientos')).filter(rel_count=0)
    
    # Construir matriz de trazabilidad
    matriz = []
    for req in requerimientos:
        fila = {
            'requerimiento': req,
            'relaciones': {}
        }
        # Obtener casos relacionados desde la tabla intermedia
        relaciones_req = req.relaciones_casos.all()  # type: ignore[attr-defined]
        for rel in relaciones_req:
            fila['relaciones'][rel.caso_de_uso.id] = {
                'relacionado': True,
                'fecha': rel.fecha_vinculacion,
                'nota': rel.nota
            }
        matriz.append(fila)
    
    # Calcular métricas
    total_requerimientos = requerimientos.count()
    total_casos = casos.count()
    total_relaciones = RequerimientoCaso.objects.filter(
        requerimiento__proyecto=proyecto
    ).count()
    
    # Cobertura
    reqs_con_casos = requerimientos.annotate(
        rel_count=Count('relaciones_casos')
    ).filter(rel_count__gt=0).count()
    
    casos_con_reqs = casos.annotate(
        rel_count=Count('relaciones_requerimientos')
    ).filter(rel_count__gt=0).count()
    
    cobertura_reqs = (reqs_con_casos / total_requerimientos * 100) if total_requerimientos > 0 else 0
    cobertura_casos = (casos_con_reqs / total_casos * 100) if total_casos > 0 else 0
    
    # Requerimientos por estado
    req_por_estado = {}
    for estado, label in Requerimiento.ESTADO_CHOICES:
        count = requerimientos.filter(estado=estado).count()
        req_por_estado[estado] = {
            'label': label,
            'count': count,
            'porcentaje': (count / total_requerimientos * 100) if total_requerimientos > 0 else 0
        }
    
    # Requerimientos por tipo
    req_por_tipo = {}
    for tipo, label in Requerimiento.TIPO_CHOICES:
        count = requerimientos.filter(tipo=tipo).count()
        req_por_tipo[tipo] = {
            'label': label,
            'count': count,
            'porcentaje': (count / total_requerimientos * 100) if total_requerimientos > 0 else 0
        }
    
    # Huérfanos
    reqs_huerfanos = requerimientos.annotate(
        rel_count=Count('relaciones_casos')
    ).filter(rel_count=0)
    
    casos_huerfanos = casos.annotate(
        rel_count=Count('relaciones_requerimientos')
    ).filter(rel_count=0)
    
    context = {
        'proyecto': proyecto,
        'requerimientos': requerimientos,
        'casos': casos,
        'matriz': matriz,
        'total_requerimientos': total_requerimientos,
        'total_casos': total_casos,
        'total_relaciones': total_relaciones,
        'reqs_con_casos': reqs_con_casos,
        'casos_con_reqs': casos_con_reqs,
        'cobertura_reqs': round(cobertura_reqs, 1),
        'cobertura_casos': round(cobertura_casos, 1),
        'req_por_estado': req_por_estado,
        'req_por_tipo': req_por_tipo,
        'reqs_huerfanos': reqs_huerfanos,
        'casos_huerfanos': casos_huerfanos,
        'es_lider': es_lider,
        'page_title': f'{proyecto.nombre} - Matriz de Trazabilidad',
        # Filtros actuales
        'filtro_tipo_req': tipo_req,
        'filtro_estado_req': estado_req,
        'filtro_solo_huerfanos': solo_huerfanos,
        'filtro_solo_sin_cubrir': solo_sin_cubrir,
    }
    
    return render(request, 'proyectos/matriz_trazabilidad.html', context)


@login_required
def exportar_matriz(request, proyecto_id, formato):
    """
    Exporta la matriz de trazabilidad en diferentes formatos: PDF, CSV.
    """
    from requerimientos.models import Requerimiento, RequerimientoCaso
    from casos_de_uso.models import CasoDeUso
    from django.http import HttpResponse
    import csv
    from datetime import datetime
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Verificar permisos
    es_lider = proyecto.lider == request.user
    es_participante = proyecto.participantes.filter(id=request.user.id).exists()
    
    if not (es_lider or es_participante):
        messages.error(request, "No tienes permiso para exportar la matriz de este proyecto.")
        return redirect("dashboards:lider_dashboard")
    
    # Obtener datos
    # Solo se exportan requerimientos VALIDADOS o en estados posteriores
    requerimientos = Requerimiento.objects.filter(proyecto=proyecto).exclude(
        estado='BORRADOR'
    ).prefetch_related(
        'relaciones_casos__caso_de_uso'
    ).order_by('id')
    
    casos = CasoDeUso.objects.filter(proyecto=proyecto).order_by('id')
    
    if formato == 'csv':
        # Exportar a CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="matriz_trazabilidad_{proyecto.nombre}_{datetime.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        
        # Encabezado
        encabezado = ['Requerimiento', 'Tipo', 'Estado']
        for caso in casos:
            encabezado.append(f'CU-{caso.pk}: {caso.nombre}')
        writer.writerow(encabezado)
        
        # Datos
        for req in requerimientos:
            fila = [
                f'REQ-{req.pk}: {req.nombre}',
                req.get_tipo_display(),  # type: ignore[attr-defined]
                req.get_estado_display()  # type: ignore[attr-defined]
            ]
            
            # Verificar relaciones con cada caso
            casos_relacionados_ids = set(
                req.relaciones_casos.values_list('caso_de_uso_id', flat=True)  # type: ignore[attr-defined]
            )
            
            for caso in casos:
                if caso.pk in casos_relacionados_ids:
                    fila.append('✓')
                else:
                    fila.append('')
            
            writer.writerow(fila)
        
        return response
    
    elif formato == 'excel':
        # Exportar a Excel (requiere openpyxl)
        try:
            from openpyxl import Workbook  # type: ignore[import-untyped]
            from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import-untyped]
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Matriz de Trazabilidad"  # type: ignore[union-attr]
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            check_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            # Encabezado
            encabezado = ['Requerimiento', 'Tipo', 'Estado']
            for caso in casos:
                encabezado.append(f'CU-{caso.pk}')
            
            ws.append(encabezado)  # type: ignore[union-attr]
            
            # Estilo del encabezado
            for cell in ws[1]:  # type: ignore[index]
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            for req in requerimientos:
                casos_relacionados_ids = set(
                    req.relaciones_casos.values_list('caso_de_uso_id', flat=True)  # type: ignore[attr-defined]
                )
                
                fila = [
                    f'REQ-{req.pk}: {req.nombre}',
                    req.get_tipo_display(),  # type: ignore[attr-defined]
                    req.get_estado_display()  # type: ignore[attr-defined]
                ]
                
                for caso in casos:
                    if caso.pk in casos_relacionados_ids:
                        fila.append('✓')
                    else:
                        fila.append('')
                
                ws.append(fila)  # type: ignore[union-attr]
                
                # Estilo para checkmarks
                row_num = ws.max_row  # type: ignore[union-attr]
                for col_num in range(4, len(encabezado) + 1):
                    cell = ws.cell(row=row_num, column=col_num)  # type: ignore[union-attr]
                    if cell.value == '✓':
                        cell.fill = check_fill
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 50  # type: ignore[union-attr, index]
            ws.column_dimensions['B'].width = 15  # type: ignore[union-attr, index]
            ws.column_dimensions['C'].width = 15  # type: ignore[union-attr, index]
            for col_num in range(4, len(encabezado) + 1):
                ws.column_dimensions[chr(64 + col_num)].width = 12  # type: ignore[union-attr, index]
            
            # Guardar en response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="matriz_trazabilidad_{proyecto.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            
            return response
            
        except ImportError:
            messages.error(request, "La exportación a Excel requiere instalar 'openpyxl'. Usa CSV como alternativa.")
            return redirect('proyectos:matriz_trazabilidad', proyecto_id=proyecto_id)
    
    elif formato == 'pdf':
        # Exportar a PDF (requiere reportlab)
        try:
            from reportlab.lib import colors  # type: ignore[import-untyped]
            from reportlab.lib.pagesizes import A4, landscape  # type: ignore[import-untyped]
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image  # type: ignore[import-untyped]
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore[import-untyped]
            from reportlab.lib.units import cm, inch  # type: ignore[import-untyped]
            from reportlab.lib.enums import TA_CENTER, TA_LEFT  # type: ignore[import-untyped]
            from io import BytesIO
            from django.db.models import Count
            from django.conf import settings
            import os
            
            # Calcular métricas del proyecto
            total_reqs = requerimientos.count()
            total_casos = casos.count()
            
            # Requerimientos por estado
            reqs_por_estado = requerimientos.values('estado').annotate(count=Count('id'))
            estado_counts = {item['estado']: item['count'] for item in reqs_por_estado}
            
            # Trazabilidad
            reqs_con_casos = requerimientos.annotate(casos_count=Count('relaciones_casos')).filter(casos_count__gt=0).count()
            casos_con_reqs = casos.annotate(reqs_count=Count('relaciones_requerimientos')).filter(reqs_count__gt=0).count()
            
            reqs_huerfanos = total_reqs - reqs_con_casos
            casos_huerfanos = total_casos - casos_con_reqs
            
            cobertura_reqs = round((reqs_con_casos / total_reqs * 100), 1) if total_reqs > 0 else 0
            cobertura_casos = round((casos_con_reqs / total_casos * 100), 1) if total_casos > 0 else 0
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm)
            elements = []
            styles = getSampleStyleSheet()
            
            # Estilos personalizados
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=18,
                spaceAfter=20,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#2c3e50')
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=15,
                textColor=colors.HexColor('#34495e')
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=8
            )
            
            # === ENCABEZADO - LOGO DEL SISTEMA ===
            try:
                from reportlab.platypus import Image, Table, TableStyle
                import os

                # Logo GRCU en el encabezado (centrado)
                logo_path = os.path.join(settings.BASE_DIR, 'accounts', 'static', 'accounts', 'img', 'logo_grcu_manager.png')
                if os.path.exists(logo_path):
                    # Logo GRCU: 1920x544px → relación de aspecto = 1920/544 ≈ 3.53
                    # Para height=2.5cm, width = 2.5cm * 3.53 ≈ 8.825cm
                    logo_grcu = Image(logo_path, width=8.825*cm, height=2.5*cm)
                    logo_grcu.hAlign = 'CENTER'
                    elements.append(logo_grcu)
                    elements.append(Spacer(1, 0.5*cm))

            except Exception as e:
                # Si hay error con el logo, continuar sin él
                pass

            # Título principal centrado
            main_title = Paragraph("<b>GRCU Manager</b><br/>Sistema de Gestión de Requerimientos", title_style)
            elements.append(main_title)
            elements.append(Spacer(1, 0.5*cm))
            
            # Información del proyecto (sin líder, solo grupo)
            project_info = []
            
            project_info.append(Paragraph(f"""
            <b>Proyecto:</b> {proyecto.nombre}<br/>
            <b>Grupo:</b> {proyecto.grupo.nombre if hasattr(proyecto, 'grupo') and proyecto.grupo else 'Sin grupo asignado'}<br/>
            <b>Fecha de Generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
            <b>Metodología:</b> {proyecto.get_metodologia_display() if hasattr(proyecto, 'metodologia') else 'No definida'}
            """, normal_style))
            
            for element in project_info:
                elements.append(element)
            
            elements.append(Spacer(1, 0.5*cm))
            
            # === DESCRIPCIÓN Y PROPÓSITO ===
            elements.append(Paragraph("<b>DESCRIPCIÓN Y PROPÓSITO</b>", subtitle_style))
            
            # Descripción del proyecto
            if proyecto.descripcion and proyecto.descripcion.strip():
                descripcion_texto = proyecto.descripcion
            else:
                descripcion_texto = "Este proyecto no tiene una descripción definida."
            
            elements.append(Paragraph(descripcion_texto, normal_style))
            elements.append(Spacer(1, 0.5*cm))
            
            # === EQUIPO DEL PROYECTO ===
            elements.append(Paragraph("<b>EQUIPO DEL PROYECTO</b>", subtitle_style))
            
            # Obtener participaciones del proyecto
            participaciones = ParticipacionProyecto.objects.filter(proyecto=proyecto).select_related('usuario', 'rol').order_by('rol__nombre', 'usuario__nombre')
            
            if participaciones.exists():
                team_data = []
                for participacion in participaciones:
                    usuario = participacion.usuario
                    rol_nombre = participacion.rol.nombre if participacion.rol else "Sin rol"
                    
                    # Intentar cargar avatar desde URL (Google OAuth)
                    avatar_cell = None
                    if usuario.avatar:
                        try:
                            import requests
                            from io import BytesIO
                            
                            # Descargar imagen desde URL
                            response = requests.get(usuario.avatar, timeout=5)
                            if response.status_code == 200:
                                # Crear imagen desde bytes
                                img_data = BytesIO(response.content)
                                avatar_cell = Image(img_data, width=1.5*cm, height=1.5*cm)
                        except Exception as e:
                            # Si falla la descarga, continuar sin avatar
                            pass
                    
                    # Si no hay avatar o falló la descarga, usar placeholder
                    if not avatar_cell:
                        avatar_cell = Paragraph("👤", ParagraphStyle('AvatarPlaceholder', fontSize=20, alignment=TA_CENTER))
                    
                    # Nombre y rol
                    nombre_texto = Paragraph(f"<b>{usuario.nombre}</b><br/><i>{rol_nombre}</i>", 
                                            ParagraphStyle('TeamMember', parent=styles['Normal'], fontSize=9, alignment=TA_LEFT))
                    
                    team_data.append([avatar_cell, nombre_texto])
                
                # Crear tabla de equipo (2 columnas: avatar, nombre+rol)
                team_table = Table(team_data, colWidths=[2*cm, 14*cm])
                team_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(team_table)
            else:
                elements.append(Paragraph("<i>No hay participantes asignados a este proyecto.</i>", normal_style))
            
            elements.append(Spacer(1, 1*cm))
            
            # === RESUMEN EJECUTIVO ===
            elements.append(Paragraph("<b>RESUMEN EJECUTIVO</b>", subtitle_style))
            
            summary_data = [
                ['Métrica', 'Valor', 'Descripción'],
                ['Total de Requerimientos', str(total_reqs), 'Requerimientos registrados en el proyecto'],
                ['Total de Casos de Uso', str(total_casos), 'Casos de uso definidos'],
                ['Requerimientos Borrador', str(estado_counts.get('BORRADOR', 0)), 'Pendientes de validación'],
                ['Requerimientos en Progreso', str(estado_counts.get('EN_PROGRESO', 0)), 'En desarrollo activo'],
                ['Requerimientos Validados', str(estado_counts.get('VALIDADO', 0)), 'Aprobados para implementación'],
                ['Requerimientos Completados', str(estado_counts.get('COMPLETADO', 0)), 'Finalizados'],
                ['Cobertura de Requerimientos', f'{cobertura_reqs}%', f'{reqs_con_casos}/{total_reqs} con casos de uso'],
                ['Cobertura de Casos de Uso', f'{cobertura_casos}%', f'{casos_con_reqs}/{total_casos} con requerimientos'],
                ['Requerimientos Huérfanos', str(reqs_huerfanos), 'Sin casos de uso asociados'],
                ['Casos de Uso Huérfanos', str(casos_huerfanos), 'Sin requerimientos asociados']
            ]
            
            # Anchos ajustados: Métrica (5.5cm) + Valor (2cm) + Descripción (9cm) = 16.5cm total
            summary_table = Table(summary_data, colWidths=[5.5*cm, 2*cm, 9*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 1*cm))
            
            # === MATRIZ DE TRAZABILIDAD ===
            elements.append(Paragraph("<b>MATRIZ DE TRAZABILIDAD</b>", subtitle_style))
            elements.append(Paragraph("Relación entre Requerimientos y Casos de Uso", normal_style))
            elements.append(Spacer(1, 0.5*cm))
            
            # Construir tabla de matriz
            data = []
            
            # Encabezado
            encabezado = ['Requerimiento', 'Estado']
            for caso in casos:
                encabezado.append(f'CU-{caso.pk}')
            data.append(encabezado)
            
            # Datos
            for req in requerimientos:
                casos_relacionados_ids = set(
                    req.relaciones_casos.values_list('caso_de_uso_id', flat=True)  # type: ignore[attr-defined]
                )
                
                # Verificar si el requerimiento es huérfano
                es_req_huerfano = len(casos_relacionados_ids) == 0
                
                fila = [
                    f'REQ-{req.pk}\n{req.nombre}',
                    req.get_estado_display()  # type: ignore[attr-defined]
                ]
                
                for caso in casos:
                    # Verificar si el caso es huérfano
                    casos_del_caso = caso.relaciones_requerimientos.values_list('requerimiento_id', flat=True)  # type: ignore[attr-defined]
                    es_caso_huerfano = len(casos_del_caso) == 0
                    
                    if caso.pk in casos_relacionados_ids:
                        fila.append('✓')
                    elif es_req_huerfano or es_caso_huerfano:
                        fila.append('⚠')  # Indicador de huérfano
                    else:
                        fila.append('')
                
                data.append(fila)
            
            # Crear tabla con mejor formato
            col_widths = [3*cm, 2*cm] + [1.2*cm] * len(casos)
            table = Table(data, colWidths=col_widths)
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                
                # Special styling for checkmarks (green)
                ('BACKGROUND', (2, 1), (-1, -1), colors.HexColor('#d5f4e6')),
                ('TEXTCOLOR', (2, 1), (-1, -1), colors.HexColor('#27ae60')),
                ('FONTNAME', (2, 1), (-1, -1), 'Helvetica-Bold'),
                
                # Special styling for warnings (red/orange)
                ('BACKGROUND', (2, 1), (-1, -1), colors.HexColor('#ffebee')),
                ('TEXTCOLOR', (2, 1), (-1, -1), colors.HexColor('#e74c3c')),
                ('FONTNAME', (2, 1), (-1, -1), 'Helvetica-Bold'),
            ]))
            
            # Aplicar estilos específicos para celdas huérfanas
            for row_idx, req in enumerate(requerimientos, 1):
                casos_relacionados_ids = set(
                    req.relaciones_casos.values_list('caso_de_uso_id', flat=True)  # type: ignore[attr-defined]
                )
                es_req_huerfano = len(casos_relacionados_ids) == 0
                
                for col_idx, caso in enumerate(casos, 2):
                    casos_del_caso = caso.relaciones_requerimientos.values_list('requerimiento_id', flat=True)  # type: ignore[attr-defined]
                    es_caso_huerfano = len(casos_del_caso) == 0
                    
                    if caso.pk in casos_relacionados_ids:
                        # Verde para relaciones existentes
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#d5f4e6')),
                            ('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#27ae60')),
                        ]))
                    elif es_req_huerfano or es_caso_huerfano:
                        # Rojo para huérfanos
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#ffebee')),
                            ('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#e74c3c')),
                        ]))
            
            elements.append(table)
            elements.append(Spacer(1, 1*cm))
            
            # === RECOMENDACIONES ===
            elements.append(Paragraph("<b>RECOMENDACIONES</b>", subtitle_style))
            
            recomendaciones = []
            
            if cobertura_reqs < 80:
                recomendaciones.append("• Mejorar la cobertura de requerimientos vinculando más casos de uso")
            if cobertura_casos < 80:
                recomendaciones.append("• Aumentar la cobertura de casos de uso asignando más requerimientos")
            if reqs_huerfanos > 0:
                recomendaciones.append(f"• Revisar {reqs_huerfanos} requerimiento(s) huérfano(s) sin casos de uso asociados")
            if casos_huerfanos > 0:
                recomendaciones.append(f"• Revisar {casos_huerfanos} caso(s) de uso huérfano(s) sin requerimientos asociados")
            if estado_counts.get('BORRADOR', 0) > total_reqs * 0.3:
                recomendaciones.append("• Avanzar los requerimientos en estado 'Borrador' hacia validación")
            
            if not recomendaciones:
                recomendaciones.append("• El proyecto muestra una buena trazabilidad general")
            
            for rec in recomendaciones:
                elements.append(Paragraph(rec, normal_style))
            
            elements.append(Spacer(1, 1*cm))
            
            # === PIE DE PÁGINA CON LOGOS ===
            try:
                from PIL import Image as PILImage
                
                # Logo del proyecto (izquierda) - calcular proporción y escalar
                logo_proyecto = None
                if proyecto.logo:
                    proyecto_logo_path = os.path.join(settings.MEDIA_ROOT, str(proyecto.logo))
                    if os.path.exists(proyecto_logo_path):
                        # Leer dimensiones originales
                        with PILImage.open(proyecto_logo_path) as img:
                            orig_width, orig_height = img.size
                            aspect_ratio = orig_width / orig_height
                            
                            # Altura deseada 2.5cm, calcular ancho manteniendo proporción
                            target_height = 2.5 * cm
                            target_width = target_height * aspect_ratio
                            
                            logo_proyecto = Image(proyecto_logo_path, width=target_width, height=target_height)
                            logo_proyecto.hAlign = 'LEFT'

                # Logo del grupo (derecha) - calcular proporción y escalar
                logo_grupo = None
                if proyecto.grupo and proyecto.grupo.logo:
                    grupo_logo_path = os.path.join(settings.MEDIA_ROOT, str(proyecto.grupo.logo))
                    if os.path.exists(grupo_logo_path):
                        # Leer dimensiones originales
                        with PILImage.open(grupo_logo_path) as img:
                            orig_width, orig_height = img.size
                            aspect_ratio = orig_width / orig_height
                            
                            # Altura deseada 2.5cm, calcular ancho manteniendo proporción
                            target_height = 2.5 * cm
                            target_width = target_height * aspect_ratio
                            
                            logo_grupo = Image(grupo_logo_path, width=target_width, height=target_height)
                            logo_grupo.hAlign = 'RIGHT'

                # Crear tabla con logos y nombres en el pie
                if logo_proyecto or logo_grupo:
                    # Estilo para los nombres
                    nombre_style = ParagraphStyle(
                        'NombreLogo',
                        parent=styles['Normal'],
                        fontSize=8,
                        textColor=colors.HexColor('#2c3e50'),
                        alignment=TA_CENTER,
                        spaceAfter=0
                    )
                    
                    # Celda izquierda: logo y nombre del proyecto
                    left_content = []
                    if logo_proyecto:
                        left_content.append(logo_proyecto)
                        left_content.append(Spacer(1, 0.1*cm))
                        left_content.append(Paragraph(f"<b>{proyecto.nombre}</b>", nombre_style))
                    
                    # Celda derecha: logo y nombre del grupo
                    right_content = []
                    if logo_grupo:
                        right_content.append(logo_grupo)
                        right_content.append(Spacer(1, 0.1*cm))
                        right_content.append(Paragraph(f"<b>{proyecto.grupo.nombre}</b>", nombre_style))
                    
                    # Crear sub-tablas para cada celda (para apilar logo + nombre)
                    left_cell = left_content if left_content else ""
                    right_cell = right_content if right_content else ""
                    
                    logos_footer_table = Table([[left_cell, right_cell]], colWidths=[9*cm, 9*cm])
                    logos_footer_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ]))
                    elements.append(logos_footer_table)
                    elements.append(Spacer(1, 0.3*cm))
            except Exception as e:
                pass

            footer_text = Paragraph("""
            <b>GRCU Manager</b> - Universidad Nacional de la Patagonia Austral<br/>
            <i>Reporte generado automáticamente • Sistema de Gestión de Requerimientos y Casos de Uso</i>
            """, ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey))
            elements.append(footer_text)
            
            doc.build(elements)
            
            # Retornar PDF
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="reporte_proyecto_{proyecto.nombre}_{datetime.now().strftime("%Y%m%d")}.pdf"'
            
            return response
            
        except ImportError:
            messages.error(request, "La exportación a PDF requiere instalar 'reportlab'. Usa CSV como alternativa.")
            return redirect('proyectos:matriz_trazabilidad', proyecto_id=proyecto_id)
    
    else:
        messages.error(request, "Formato de exportación no válido.")
        return redirect('proyectos:matriz_trazabilidad', proyecto_id=proyecto_id)


@login_required
def proyecto_reportes(request, proyecto_id):
    """
    Vista de reportes y estadísticas del proyecto.
    Muestra métricas, gráficos y análisis del estado del proyecto.
    Accesible por líder y participantes del proyecto.
    """
    from requerimientos.models import Requerimiento
    from casos_de_uso.models import CasoDeUso
    from django.db.models import Count, Q
    import json
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Verificar permisos: líder o participante del proyecto
    es_lider = proyecto.lider == request.user
    es_participante = proyecto.participantes.filter(id=request.user.id).exists()
    
    if not (es_lider or es_participante):
        messages.error(request, "No tienes permiso para ver los reportes de este proyecto.")
        return redirect("dashboards:lider_dashboard")
    
    # === MÉTRICAS GENERALES ===
    total_requerimientos = Requerimiento.objects.filter(proyecto=proyecto).count()
    total_casos_uso = CasoDeUso.objects.filter(proyecto=proyecto).count()
    
    # === REQUERIMIENTOS POR TIPO ===
    reqs_funcionales = Requerimiento.objects.filter(proyecto=proyecto, tipo='FUNCIONAL').count()
    reqs_no_funcionales = Requerimiento.objects.filter(proyecto=proyecto, tipo='NO_FUNCIONAL').count()
    
    # === REQUERIMIENTOS POR ESTADO ===
    reqs_pendientes = Requerimiento.objects.filter(proyecto=proyecto, estado='PENDIENTE').count()
    reqs_en_desarrollo = Requerimiento.objects.filter(proyecto=proyecto, estado='EN_DESARROLLO').count()
    reqs_aprobados = Requerimiento.objects.filter(proyecto=proyecto, estado='APROBADO').count()
    
    # === TRAZABILIDAD ===
    reqs_con_casos = Requerimiento.objects.filter(proyecto=proyecto).annotate(
        casos_count=Count('casos_relacionados')
    ).filter(casos_count__gt=0).count()
    
    reqs_huerfanos = total_requerimientos - reqs_con_casos
    
    casos_con_reqs = CasoDeUso.objects.filter(proyecto=proyecto).annotate(
        reqs_count=Count('requerimientos_relacionados')
    ).filter(reqs_count__gt=0).count()
    
    casos_huerfanos = total_casos_uso - casos_con_reqs
    
    # === COBERTURA ===
    cobertura_reqs = (reqs_con_casos / total_requerimientos * 100) if total_requerimientos > 0 else 0
    cobertura_casos = (casos_con_reqs / total_casos_uso * 100) if total_casos_uso > 0 else 0
    
    # === DATOS PARA GRÁFICOS (JSON) ===
    # Gráfico de Requerimientos por Tipo
    tipo_labels = ['Funcionales', 'No Funcionales']
    tipo_values = [reqs_funcionales, reqs_no_funcionales]
    tipo_colors = ['#3498db', '#e74c3c']
    
    # Gráfico de Requerimientos por Estado
    estado_labels = ['Pendiente', 'En Desarrollo', 'Aprobado']
    estado_values = [reqs_pendientes, reqs_en_desarrollo, reqs_aprobados]
    estado_colors = ['#f39c12', '#3498db', '#27ae60']
    
    # Gráfico de Trazabilidad
    trazabilidad_labels = ['Con Casos de Uso', 'Huérfanos']
    trazabilidad_reqs_values = [reqs_con_casos, reqs_huerfanos]
    trazabilidad_casos_values = [casos_con_reqs, casos_huerfanos]
    
    context = {
        'proyecto': proyecto,
        'es_lider': es_lider,
        'page_title': f'{proyecto.nombre} - Reportes',
        
        # Métricas
        'total_requerimientos': total_requerimientos,
        'total_casos_uso': total_casos_uso,
        'reqs_funcionales': reqs_funcionales,
        'reqs_no_funcionales': reqs_no_funcionales,
        'reqs_pendientes': reqs_pendientes,
        'reqs_en_desarrollo': reqs_en_desarrollo,
        'reqs_aprobados': reqs_aprobados,
        'reqs_con_casos': reqs_con_casos,
        'reqs_huerfanos': reqs_huerfanos,
        'casos_con_reqs': casos_con_reqs,
        'casos_huerfanos': casos_huerfanos,
        'cobertura_reqs': round(cobertura_reqs, 1),
        'cobertura_casos': round(cobertura_casos, 1),
        
        # Datos para gráficos (JSON)
        'tipo_labels_json': json.dumps(tipo_labels),
        'tipo_values_json': json.dumps(tipo_values),
        'tipo_colors_json': json.dumps(tipo_colors),
        'estado_labels_json': json.dumps(estado_labels),
        'estado_values_json': json.dumps(estado_values),
        'estado_colors_json': json.dumps(estado_colors),
        'trazabilidad_labels_json': json.dumps(trazabilidad_labels),
        'trazabilidad_reqs_values_json': json.dumps(trazabilidad_reqs_values),
        'trazabilidad_casos_values_json': json.dumps(trazabilidad_casos_values),
    }
    
    return render(request, 'proyectos/proyecto_reportes.html', context)


@login_required
def gestionar_integrantes(request, proyecto_id):
    """
    Vista para que el líder del proyecto gestione los roles de los integrantes.
    Solo puede asignar roles de: Desarrollador, Stakeholder (Cliente) y Visitante.
    """
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Validar que el usuario es el líder del proyecto
    if proyecto.lider != request.user:
        messages.error(request, "No tienes permisos para gestionar los integrantes de este proyecto.")
        return redirect('dashboards:lider_dashboard')
    
    # Obtener roles permitidos para asignar
    rol_desarrollador = get_object_or_404(Rol, nombre="Desarrollador")
    rol_stakeholder = get_object_or_404(Rol, nombre="Stakeholder")
    rol_visitante = get_object_or_404(Rol, nombre="Visitante")
    
    roles_permitidos = [rol_desarrollador, rol_stakeholder, rol_visitante]
    
    # Si es POST, procesar cambios de roles
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('rol_'):
                try:
                    usuario_id = int(key.split('_')[1])
                    nuevo_rol_id = int(value)
                    
                    # Validar que el usuario es participante del proyecto
                    usuario = proyecto.participantes.get(id=usuario_id)
                    
                    # Validar que el rol es permitido
                    nuevo_rol = Rol.objects.get(id=nuevo_rol_id)
                    if nuevo_rol not in roles_permitidos:
                        messages.error(request, f"El rol {nuevo_rol.nombre} no está permitido.")
                        continue
                    
                    # No se puede cambiar el rol del líder del proyecto
                    if usuario == proyecto.lider:
                        messages.warning(request, f"No puedes cambiar tu propio rol como líder del proyecto.")
                        continue
                    
                    # Actualizar o crear la participación
                    participacion, created = ParticipacionProyecto.objects.update_or_create(
                        usuario=usuario,
                        proyecto=proyecto,
                        defaults={'rol': nuevo_rol}
                    )
                    
                except (ValueError, Usuario.DoesNotExist, Rol.DoesNotExist):
                    continue
        
        messages.success(request, "Los roles de los integrantes han sido actualizados correctamente.")
        return redirect('proyectos:gestionar_integrantes', proyecto_id=proyecto_id)
    
    # Obtener participaciones actuales
    participaciones = ParticipacionProyecto.objects.filter(proyecto=proyecto).select_related('usuario', 'rol')
    
    # Crear un diccionario de usuarios con sus roles actuales
    usuarios_con_roles = []
    for participante in proyecto.participantes.all():
        try:
            participacion = participaciones.get(usuario=participante)
            rol_actual = participacion.rol
        except ParticipacionProyecto.DoesNotExist:
            # Si no tiene participación, asignar rol por defecto (Desarrollador)
            rol_actual = rol_desarrollador
            ParticipacionProyecto.objects.create(
                usuario=participante,
                proyecto=proyecto,
                rol=rol_desarrollador
            )
        
        usuarios_con_roles.append({
            'usuario': participante,
            'rol_actual': rol_actual,
            'es_lider': participante == proyecto.lider,
        })
    
    context = {
        'page_title': f'{proyecto.nombre} - Gestión de Integrantes',
        'proyecto': proyecto,
        'usuarios_con_roles': usuarios_con_roles,
        'roles_permitidos': roles_permitidos,
    }
    
    return render(request, 'proyectos/gestionar_integrantes.html', context)


@login_required
def proyecto_detail_admin(request, proyecto_id):
    """
    Vista detallada de un proyecto para administradores.
    Muestra toda la información del proyecto incluyendo integrantes, requerimientos,
    casos de uso, métricas y gráficos.
    """
    from requerimientos.models import Requerimiento, RequerimientoCaso
    from casos_de_uso.models import CasoDeUso
    from auditoria.models import RegistroActividad
    from django.db.models import Count
    import json
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    integrantes = list(proyecto.participantes.all())
    lider = proyecto.lider
    requerimientos = Requerimiento.objects.filter(proyecto=proyecto)
    casos = CasoDeUso.objects.filter(proyecto=proyecto)
    acciones = RegistroActividad.objects.filter(usuario__in=integrantes).order_by('-fecha')[:5]  # Solo 5 acciones más recientes
    
    # Huérfanos definidos como aquellos sin relación persistida en la tabla intermedia RequerimientoCaso
    reqs_huerfanos = requerimientos.annotate(rel_count=Count('relaciones_casos')).filter(rel_count=0)
    casos_huerfanos = casos.annotate(rel_count=Count('relaciones_requerimientos')).filter(rel_count=0)
    reqs_huerfanos_ids = list(reqs_huerfanos.values_list('pk', flat=True))
    casos_huerfanos_ids = list(casos_huerfanos.values_list('pk', flat=True))
    
    # Matriz de trazabilidad simple: relacionar requerimientos y casos por nombre parcial (heurística)
    matriz = []
    for req in requerimientos:
        relacionados = [cu for cu in casos if req.nombre.split()[0].lower() in cu.nombre.lower() or req.nombre.lower() in cu.descripcion.lower()]
        matriz.append({'req': req, 'casos': relacionados})
    
    # Agregaciones para gráficos
    # Requerimientos por estado - Usar los estados correctos del modelo
    req_estado_qs = requerimientos.values('estado').annotate(count=Count('id'))
    req_estado_map = {item['estado']: item['count'] for item in req_estado_qs}
    req_estado_labels = ["Borrador", "Validado", "Priorizado", "En Proceso", "Terminado"]
    req_estado_values = [
        req_estado_map.get('BORRADOR', 0),
        req_estado_map.get('VALIDADO', 0),
        req_estado_map.get('PRIORIZADO', 0),
        req_estado_map.get('EN_PROCESO', 0),
        req_estado_map.get('TERMINADO', 0)
    ]
    
    # Requerimientos por tipo - Usar los tipos correctos del modelo
    req_tipo_qs = requerimientos.values('tipo').annotate(count=Count('id'))
    req_tipo_map = {item['tipo']: item['count'] for item in req_tipo_qs}
    req_tipo_labels = ["Funcional", "No Funcional", "Sistema"]
    req_tipo_values = [
        req_tipo_map.get('FUNCIONAL', 0),
        req_tipo_map.get('NO_FUNCIONAL', 0),
        req_tipo_map.get('SISTEMA', 0)
    ]
    
    # Casos de uso: conteo por disponibilidad de detalle (Tradicional / Ágil / Sin detalle)
    casos_trad = casos.filter(detalle_tradicional_reverse__isnull=False).count()
    casos_agil = casos.filter(detalle_agil_reverse__isnull=False).count()
    casos_sin = casos.filter(detalle_agil_reverse__isnull=True, detalle_tradicional_reverse__isnull=True).count()
    casos_tipo_labels = ["Tradicional", "Ágil", "Sin detalle"]
    casos_tipo_values = [casos_trad, casos_agil, casos_sin]
    
    # Acciones por usuario (top 5) - Limitar a 5 acciones en el timeline también
    acciones_por_usuario_qs = RegistroActividad.objects.filter(usuario__in=integrantes).values('usuario__nombre').annotate(count=Count('id')).order_by('-count')[:5]
    acciones_labels = [a['usuario__nombre'] for a in acciones_por_usuario_qs]
    acciones_values = [a['count'] for a in acciones_por_usuario_qs]
    
    context = {
        'page_title': f'{proyecto.nombre} - Detalle del Proyecto',
        'proyecto': proyecto,
        'integrantes': integrantes,
        'lider': lider,
        'requerimientos': requerimientos,
        'casos': casos,
        'acciones': acciones,
        'reqs_huerfanos': reqs_huerfanos,
        'reqs_huerfanos_ids': reqs_huerfanos_ids,
        'casos_huerfanos': casos_huerfanos,
        'casos_huerfanos_ids': casos_huerfanos_ids,
        'matriz': matriz,
        # Datos para gráficos (convertidos a JSON para Chart.js)
        'req_estado_labels': json.dumps(req_estado_labels),
        'req_estado_values': json.dumps(req_estado_values),
        'req_tipo_labels': json.dumps(req_tipo_labels),
        'req_tipo_values': json.dumps(req_tipo_values),
        'casos_tipo_labels': json.dumps(casos_tipo_labels),
        'casos_tipo_values': json.dumps(casos_tipo_values),
        'acciones_labels': json.dumps(acciones_labels),
        'acciones_values': json.dumps(acciones_values),
    }
    
    return render(request, 'proyectos/proyecto_detail_admin.html', context)


def crear_header_footer(canvas, doc, proyecto, logo_proyecto_path, logo_grupo_path, total_pages=0):
    """
    Función para agregar encabezado y pie de página en todas las páginas del reporte.
    
    Args:
        canvas: Canvas de ReportLab
        doc: Documento de ReportLab
        proyecto: Instancia del proyecto
        logo_proyecto_path: Ruta al logo del proyecto
        logo_grupo_path: Ruta al logo del grupo
        total_pages: Total de páginas del documento (0 si no se conoce)
    """
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from datetime import datetime
    import os
    from django.conf import settings
    
    canvas.saveState()
    width, height = doc.pagesize
    
    # === MARCA DE AGUA (si el proyecto está en estado BORRADOR) ===
    if hasattr(proyecto, 'estado') and proyecto.estado == 'BORRADOR':
        canvas.setFont('Helvetica-Bold', 60)
        canvas.setFillColor(colors.Color(0.9, 0.9, 0.9, alpha=0.3))
        canvas.saveState()
        canvas.translate(width/2, height/2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "BORRADOR")
        canvas.restoreState()
    
    # === ENCABEZADO ===
    # Logo de GRCU Manager (izquierda)
    try:
        logo_grcu_path = os.path.join(settings.BASE_DIR, 'accounts', 'static', 'accounts', 'img', 'logo_grcu_manager.png')
        if os.path.exists(logo_grcu_path):
            canvas.drawImage(logo_grcu_path, 2*cm, height - 2.5*cm, 
                           width=2*cm, height=0.6*cm, 
                           preserveAspectRatio=True, mask='auto')
    except:
        pass
    
    # Línea superior
    canvas.setStrokeColor(colors.HexColor('#2c3e50'))
    canvas.setLineWidth(2)
    canvas.line(2*cm, height - 2*cm, width - 2*cm, height - 2*cm)
    
    # Nombre del proyecto (izquierda, línea 1) - Truncado a 40 caracteres
    canvas.setFont('Helvetica-Bold', 9)
    canvas.setFillColor(colors.HexColor('#2c3e50'))
    proyecto_nombre = proyecto.nombre[:40] + '...' if len(proyecto.nombre) > 40 else proyecto.nombre
    canvas.drawString(2*cm, height - 1.7*cm, f"Proyecto: {proyecto_nombre}")
    
    # Fecha de generación (centro, línea 1) - Más pequeña
    canvas.setFont('Helvetica', 7)
    canvas.setFillColor(colors.grey)
    fecha_texto = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    fecha_width = canvas.stringWidth(fecha_texto, 'Helvetica', 7)
    canvas.drawString((width - fecha_width) / 2, height - 1.7*cm, fecha_texto)
    
    # Número de página (derecha, línea 1)
    canvas.setFont('Helvetica-Bold', 9)
    canvas.setFillColor(colors.HexColor('#34495e'))
    if total_pages > 0:
        page_text = f"Pág {doc.page}/{total_pages}"
    else:
        page_text = f"Pág {doc.page}"
    page_width = canvas.stringWidth(page_text, 'Helvetica-Bold', 9)
    canvas.drawString(width - 2*cm - page_width, height - 1.7*cm, page_text)
    
    # Estado del proyecto como badge (línea 2 izquierda)
    if hasattr(proyecto, 'estado'):
        estado_map = {
            'PLANIFICACION': ('#3498db', 'Planificación'),
            'EN_PROGRESO': ('#f39c12', 'En Progreso'),
            'COMPLETADO': ('#27ae60', 'Completado'),
            'BORRADOR': ('#95a5a6', 'Borrador'),
        }
        color_hex, texto_estado = estado_map.get(proyecto.estado, ('#95a5a6', proyecto.estado))
        
        # Badge de estado
        badge_x = 2*cm
        badge_y = height - 2.25*cm
        badge_width = 2.2*cm
        badge_height = 0.35*cm
        
        canvas.setFillColor(colors.HexColor(color_hex))
        canvas.roundRect(badge_x, badge_y, badge_width, badge_height, 0.08*cm, fill=1, stroke=0)
        
        canvas.setFont('Helvetica-Bold', 6)
        canvas.setFillColor(colors.white)
        canvas.drawCentredString(badge_x + badge_width/2, badge_y + 0.08*cm, texto_estado)
        
        # Mini barra de progreso (junto al badge)
        try:
            from requerimientos.models import Requerimiento
            reqs = Requerimiento.objects.filter(proyecto=proyecto)
            total_reqs = reqs.count()
            
            if total_reqs > 0:
                terminados = reqs.filter(estado='TERMINADO').count()
                porcentaje = (terminados / total_reqs) * 100
                
                # Barra de progreso
                progress_x = badge_x + badge_width + 0.2*cm
                progress_y = badge_y
                progress_width = 2.5*cm
                progress_height = 0.35*cm
                
                # Fondo de la barra
                canvas.setFillColor(colors.HexColor('#ecf0f1'))
                canvas.roundRect(progress_x, progress_y, progress_width, progress_height, 0.08*cm, fill=1, stroke=0)
                
                # Barra de progreso
                if porcentaje > 0:
                    progreso_color = colors.HexColor('#27ae60') if porcentaje >= 80 else colors.HexColor('#f39c12') if porcentaje >= 50 else colors.HexColor('#e74c3c')
                    canvas.setFillColor(progreso_color)
                    canvas.roundRect(progress_x, progress_y, (progress_width * porcentaje / 100), progress_height, 0.08*cm, fill=1, stroke=0)
                
                # Texto de porcentaje
                canvas.setFont('Helvetica-Bold', 6)
                canvas.setFillColor(colors.HexColor('#2c3e50'))
                canvas.drawCentredString(progress_x + progress_width/2, progress_y + 0.08*cm, f"{porcentaje:.0f}%")
        except:
            pass
    
    # === PIE DE PÁGINA ===
    # Línea inferior
    canvas.setStrokeColor(colors.HexColor('#2c3e50'))
    canvas.setLineWidth(1)
    canvas.line(2*cm, 2*cm, width - 2*cm, 2*cm)
    
    # Logos (izquierda)
    y_logo = 0.7*cm
    x_logo = 2*cm
    logo_height = 1*cm
    
    try:
        if logo_proyecto_path and os.path.exists(logo_proyecto_path):
            canvas.drawImage(logo_proyecto_path, x_logo, y_logo, 
                           width=1.5*cm, height=logo_height, 
                           preserveAspectRatio=True, mask='auto')
            x_logo += 1.7*cm
    except:
        pass
    
    try:
        if logo_grupo_path and os.path.exists(logo_grupo_path):
            canvas.drawImage(logo_grupo_path, x_logo, y_logo, 
                           width=1.5*cm, height=logo_height, 
                           preserveAspectRatio=True, mask='auto')
    except:
        pass
    
    # Información del sistema (centro) - 3 líneas
    canvas.setFont('Helvetica', 6)
    canvas.setFillColor(colors.grey)
    sistema_texto = "GRCU Manager - Sistema de Gestión de Requerimientos"
    sistema_width = canvas.stringWidth(sistema_texto, 'Helvetica', 6)
    canvas.drawString((width - sistema_width) / 2, 1.5*cm, sistema_texto)
    
    # Grupo y versión (centro, segunda línea)
    if proyecto.grupo:
        # Truncar nombre del grupo si es muy largo
        grupo_nombre = proyecto.grupo.nombre[:30] + '...' if len(proyecto.grupo.nombre) > 30 else proyecto.grupo.nombre
        grupo_texto = f"Grupo: {grupo_nombre}"
    else:
        grupo_texto = "Sin grupo asignado"
    grupo_width = canvas.stringWidth(grupo_texto, 'Helvetica', 6)
    canvas.drawString((width - grupo_width) / 2, 1.1*cm, grupo_texto)
    
    # Universidad (centro, tercera línea)
    canvas.setFont('Helvetica-Oblique', 5)
    univ_texto = "Universidad Nacional de la Patagonia Austral - UNPA"
    univ_width = canvas.stringWidth(univ_texto, 'Helvetica-Oblique', 5)
    canvas.drawString((width - univ_width) / 2, 0.7*cm, univ_texto)
    
    # Código QR (derecha, arriba) - Reposicionado para no superponerse
    qr_size = 1.2*cm
    qr_x = width - 2*cm - qr_size
    qr_y = 0.7*cm
    
    try:
        import qrcode
        from io import BytesIO
        from reportlab.lib.utils import ImageReader
        
        # Crear URL del proyecto (ajusta según tu dominio)
        qr_url = f"http://localhost:8000/proyectos/{proyecto.pk}/"
        
        # Generar QR
        qr = qrcode.QRCode(version=1, box_size=10, border=1)
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convertir a bytes
        buffer_qr = BytesIO()
        img.save(buffer_qr, format='PNG')
        buffer_qr.seek(0)
        
        # Dibujar QR en el PDF
        canvas.drawImage(ImageReader(buffer_qr), qr_x, qr_y, 
                        width=qr_size, height=qr_size)
        
        # Texto "Escanear" debajo del QR
        canvas.setFont('Helvetica', 5)
        canvas.setFillColor(colors.grey)
        qr_texto = "Escanear"
        qr_texto_width = canvas.stringWidth(qr_texto, 'Helvetica', 5)
        canvas.drawString(qr_x + (qr_size - qr_texto_width)/2, 0.4*cm, qr_texto)
    except:
        # Si falla el QR, mostrar texto en su lugar
        pass
    
    # Información de confidencialidad (derecha, pero con espacio para el QR)
    text_right_x = qr_x - 0.3*cm  # Espacio antes del QR
    
    canvas.setFont('Helvetica-BoldOblique', 6)
    canvas.setFillColor(colors.HexColor('#e74c3c'))
    confidencial_texto = "CONFIDENCIAL"
    conf_width = canvas.stringWidth(confidencial_texto, 'Helvetica-BoldOblique', 6)
    canvas.drawRightString(text_right_x, 1.5*cm, confidencial_texto)
    
    # ID del proyecto (derecha, segunda línea)
    canvas.setFont('Helvetica', 6)
    canvas.setFillColor(colors.grey)
    id_texto = f"ID: PRY-{proyecto.pk:04d}"
    canvas.drawRightString(text_right_x, 1.1*cm, id_texto)
    
    canvas.restoreState()


@login_required
def generar_reporte_personalizado(request, proyecto_id):
    """
    Genera un reporte PDF personalizado según las secciones seleccionadas por el usuario.
    Accesible por líder, participantes y clientes del proyecto.
    """
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    
    # Verificar permisos (líder, participante o cliente)
    es_lider = proyecto.lider == request.user
    es_participante = proyecto.participantes.filter(id=request.user.id).exists()
    es_cliente = proyecto.clientes.filter(id=request.user.id).exists()
    
    if not (es_lider or es_participante or es_cliente):
        messages.error(request, "No tienes permisos para generar reportes de este proyecto.")
        return redirect('proyectos:lista_proyectos')
    
    if request.method != 'POST':
        return redirect('proyectos:proyecto_reportes', proyecto_id=proyecto_id)
    
    # Obtener secciones seleccionadas
    incluir_equipo = request.POST.get('incluir_equipo') == 'on'
    incluir_resumen = request.POST.get('incluir_resumen') == 'on'
    incluir_matriz = request.POST.get('incluir_matriz') == 'on'
    incluir_requerimientos = request.POST.get('incluir_requerimientos') == 'on'
    incluir_casos_uso = request.POST.get('incluir_casos_uso') == 'on'
    incluir_recomendaciones = request.POST.get('incluir_recomendaciones') == 'on'
    incluir_info_grupo = request.POST.get('incluir_info_grupo') == 'on'
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
        from io import BytesIO
        from django.http import HttpResponse
        from datetime import datetime
        from django.db.models import Count, Q
        import os
        from django.conf import settings
        
        # Crear buffer para el PDF
        buffer = BytesIO()
        
        # Configurar márgenes aumentados para header y footer
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=3*cm,      # Aumentado para el header con badge y progreso
            bottomMargin=2.8*cm, # Aumentado para el footer con QR
            leftMargin=2*cm,
            rightMargin=2*cm
        )
        elements = []
        
        # Obtener rutas de logos para el header/footer
        logo_proyecto_path = None
        logo_grupo_path = None
        
        if proyecto.logo:
            logo_proyecto_path = os.path.join(settings.MEDIA_ROOT, str(proyecto.logo))
        
        if proyecto.grupo and proyecto.grupo.logo:
            logo_grupo_path = os.path.join(settings.MEDIA_ROOT, str(proyecto.grupo.logo))
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            textColor=colors.HexColor('#34495e')
        )
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=8
        )
        
        # === ENCABEZADO - LOGO DEL SISTEMA ===
        try:
            from reportlab.platypus import Image
            logo_path = os.path.join(settings.BASE_DIR, 'accounts', 'static', 'accounts', 'img', 'logo_grcu_manager.png')
            if os.path.exists(logo_path):
                logo_grcu = Image(logo_path, width=8.825*cm, height=2.5*cm)
                logo_grcu.hAlign = 'CENTER'
                elements.append(logo_grcu)
                elements.append(Spacer(1, 0.5*cm))
        except Exception:
            pass
        
        # Título principal
        main_title = Paragraph("<b>GRCU Manager</b><br/>Reporte Personalizado del Proyecto", title_style)
        elements.append(main_title)
        elements.append(Spacer(1, 0.5*cm))
        
        # === INFORMACIÓN DEL PROYECTO (siempre incluida) ===
        project_info = Paragraph(f"""
        <b>Proyecto:</b> {proyecto.nombre}<br/>
        <b>Grupo:</b> {proyecto.grupo.nombre if proyecto.grupo else 'Sin grupo asignado'}<br/>
        <b>Fecha de Generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
        <b>Metodología:</b> {proyecto.get_metodologia_display()}
        """, normal_style)
        elements.append(project_info)
        elements.append(Spacer(1, 1*cm))
        
        # === TABLA DE CONTENIDOS ===
        elements.append(Paragraph("<b>ÍNDICE DE CONTENIDOS</b>", subtitle_style))
        
        toc_style = ParagraphStyle(
            'TOC',
            parent=styles['Normal'],
            fontSize=10,
            leftIndent=0.5*cm,
            spaceAfter=6,
            textColor=colors.HexColor('#2c3e50')
        )
        
        toc_items = [
            ("1.", "Descripción y Propósito", True),
        ]
        
        section_num = 2
        if incluir_equipo:
            toc_items.append((f"{section_num}.", "Equipo del Proyecto", True))
            section_num += 1
        
        if incluir_resumen:
            toc_items.append((f"{section_num}.", "Resumen Ejecutivo", True))
            section_num += 1
        
        if incluir_matriz:
            toc_items.append((f"{section_num}.", "Matriz de Trazabilidad", True))
            section_num += 1
        
        if incluir_requerimientos:
            toc_items.append((f"{section_num}.", "Listado Detallado de Requerimientos", True))
            section_num += 1
        
        if incluir_casos_uso:
            toc_items.append((f"{section_num}.", "Listado Detallado de Casos de Uso", True))
            section_num += 1
        
        if incluir_info_grupo:
            toc_items.append((f"{section_num}.", "Información del Grupo", True))
            section_num += 1
        
        if incluir_recomendaciones:
            toc_items.append((f"{section_num}.", "Recomendaciones y Análisis", True))
            section_num += 1
        
        # Renderizar tabla de contenidos
        for num, titulo, incluido in toc_items:
            if incluido:
                toc_line = f"<b>{num}</b> {titulo} {'.' * 80}"
                elements.append(Paragraph(toc_line, toc_style))
        
        elements.append(Spacer(1, 1*cm))
        elements.append(PageBreak())
        
        # === DESCRIPCIÓN Y PROPÓSITO (siempre incluida) ===
        elements.append(Paragraph("<b>DESCRIPCIÓN Y PROPÓSITO</b>", subtitle_style))
        descripcion_texto = proyecto.descripcion if proyecto.descripcion and proyecto.descripcion.strip() else "Este proyecto no tiene una descripción definida."
        elements.append(Paragraph(descripcion_texto, normal_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # === EQUIPO DEL PROYECTO (opcional) ===
        if incluir_equipo:
            elements.append(Paragraph("<b>EQUIPO DEL PROYECTO</b>", subtitle_style))
            participaciones = ParticipacionProyecto.objects.filter(proyecto=proyecto).select_related('usuario', 'rol').order_by('rol__nombre', 'usuario__nombre')
            
            if participaciones.exists():
                team_data = []
                for participacion in participaciones:
                    usuario = participacion.usuario
                    rol_nombre = participacion.rol.nombre if participacion.rol else "Sin rol"
                    
                    # Intentar cargar avatar desde URL
                    avatar_cell = None
                    if usuario.avatar:
                        try:
                            import requests
                            from io import BytesIO as ImgBytesIO
                            response = requests.get(usuario.avatar, timeout=5)
                            if response.status_code == 200:
                                img_data = ImgBytesIO(response.content)
                                avatar_cell = Image(img_data, width=1.5*cm, height=1.5*cm)
                        except Exception:
                            pass
                    
                    if not avatar_cell:
                        avatar_cell = Paragraph("👤", ParagraphStyle('AvatarPlaceholder', fontSize=20, alignment=TA_CENTER))
                    
                    nombre_texto = Paragraph(f"<b>{usuario.nombre}</b><br/><i>{rol_nombre}</i>", 
                                            ParagraphStyle('TeamMember', parent=styles['Normal'], fontSize=9, alignment=TA_LEFT))
                    team_data.append([avatar_cell, nombre_texto])
                
                team_table = Table(team_data, colWidths=[2*cm, 14*cm])
                team_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(team_table)
            else:
                elements.append(Paragraph("<i>No hay participantes asignados a este proyecto.</i>", normal_style))
            
            elements.append(Spacer(1, 1*cm))
        
        # === RESUMEN EJECUTIVO (opcional) ===
        if incluir_resumen:
            from requerimientos.models import Requerimiento
            from casos_de_uso.models import CasoDeUso
            from django.db.models import Count
            
            requerimientos = Requerimiento.objects.filter(proyecto=proyecto)
            casos = CasoDeUso.objects.filter(proyecto=proyecto)
            
            total_reqs = requerimientos.count()
            total_casos = casos.count()
            
            # Estadísticas por estado
            estado_counts = {}
            for req in requerimientos:
                estado = req.estado if hasattr(req, 'estado') else 'SIN_ESTADO'
                estado_counts[estado] = estado_counts.get(estado, 0) + 1
            
            # Trazabilidad
            reqs_con_casos = sum(1 for req in requerimientos if req.casos_relacionados.exists())
            casos_con_reqs = sum(1 for caso in casos if caso.requerimientos_relacionados.exists())
            reqs_huerfanos = total_reqs - reqs_con_casos
            casos_huerfanos = total_casos - casos_con_reqs
            
            cobertura_reqs = round((reqs_con_casos / total_reqs * 100) if total_reqs > 0 else 0, 1)
            cobertura_casos = round((casos_con_reqs / total_casos * 100) if total_casos > 0 else 0, 1)
            
            elements.append(Paragraph("<b>RESUMEN EJECUTIVO</b>", subtitle_style))
            
            summary_data = [
                ['Métrica', 'Valor', 'Descripción'],
                ['Total de Requerimientos', str(total_reqs), 'Requerimientos registrados en el proyecto'],
                ['Total de Casos de Uso', str(total_casos), 'Casos de uso definidos'],
                ['Requerimientos Borrador', str(estado_counts.get('BORRADOR', 0)), 'Pendientes de validación'],
                ['Requerimientos en Progreso', str(estado_counts.get('EN_PROGRESO', 0)), 'En desarrollo activo'],
                ['Requerimientos Validados', str(estado_counts.get('VALIDADO', 0)), 'Aprobados para implementación'],
                ['Requerimientos Completados', str(estado_counts.get('COMPLETADO', 0)), 'Finalizados'],
                ['Cobertura de Requerimientos', f'{cobertura_reqs}%', f'{reqs_con_casos}/{total_reqs} con casos de uso'],
                ['Cobertura de Casos de Uso', f'{cobertura_casos}%', f'{casos_con_reqs}/{total_casos} con requerimientos'],
                ['Requerimientos Huérfanos', str(reqs_huerfanos), 'Sin casos de uso asociados'],
                ['Casos de Uso Huérfanos', str(casos_huerfanos), 'Sin requerimientos asociados']
            ]
            
            summary_table = Table(summary_data, colWidths=[5.5*cm, 2*cm, 9*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 1*cm))
        
        # === MATRIZ DE TRAZABILIDAD (opcional) ===
        if incluir_matriz:
            from requerimientos.models import Requerimiento
            from casos_de_uso.models import CasoDeUso
            
            requerimientos = list(Requerimiento.objects.filter(proyecto=proyecto))
            casos = list(CasoDeUso.objects.filter(proyecto=proyecto))
            
            if requerimientos and casos:
                elements.append(Paragraph("<b>MATRIZ DE TRAZABILIDAD</b>", subtitle_style))
                elements.append(Paragraph("Relación entre Requerimientos y Casos de Uso", normal_style))
                elements.append(Spacer(1, 0.5*cm))
                
                # Construir tabla de matriz
                data = []
                
                # Encabezado
                encabezado = ['Requerimiento', 'Estado']
                for caso in casos:
                    encabezado.append(f'CU-{caso.pk}')
                data.append(encabezado)
                
                # Datos
                for req in requerimientos:
                    casos_relacionados_ids = set(req.relaciones_casos.values_list('caso_de_uso_id', flat=True))
                    es_req_huerfano = len(casos_relacionados_ids) == 0
                    
                    fila = [
                        f'REQ-{req.pk}\n{req.nombre}',
                        req.get_estado_display() if hasattr(req, 'get_estado_display') else 'N/A'
                    ]
                    
                    for caso in casos:
                        casos_del_caso = caso.relaciones_requerimientos.values_list('requerimiento_id', flat=True)
                        es_caso_huerfano = len(casos_del_caso) == 0
                        
                        if caso.pk in casos_relacionados_ids:
                            fila.append('✓')
                        elif es_req_huerfano or es_caso_huerfano:
                            fila.append('⚠')
                        else:
                            fila.append('')
                    
                    data.append(fila)
                
                # Crear tabla con formato mejorado
                col_widths = [3*cm, 2*cm] + [1.2*cm] * len(casos)
                table = Table(data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    # Header styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    
                    # Data rows
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                    
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                
                # Aplicar estilos específicos para celdas con colores
                for row_idx, req in enumerate(requerimientos, 1):
                    casos_relacionados_ids = set(req.relaciones_casos.values_list('caso_de_uso_id', flat=True))
                    es_req_huerfano = len(casos_relacionados_ids) == 0
                    
                    for col_idx, caso in enumerate(casos, 2):
                        casos_del_caso = caso.relaciones_requerimientos.values_list('requerimiento_id', flat=True)
                        es_caso_huerfano = len(casos_del_caso) == 0
                        
                        if caso.pk in casos_relacionados_ids:
                            # Verde para relaciones existentes
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#d5f4e6')),
                                ('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#27ae60')),
                            ]))
                        elif es_req_huerfano or es_caso_huerfano:
                            # Rojo para huérfanos
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#ffebee')),
                                ('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#e74c3c')),
                            ]))
                
                elements.append(table)
                elements.append(Spacer(1, 1*cm))
        
        # === LISTADO DE REQUERIMIENTOS (opcional) ===
        if incluir_requerimientos:
            from requerimientos.models import Requerimiento
            
            requerimientos = Requerimiento.objects.filter(proyecto=proyecto).order_by('id')
            
            if requerimientos.exists():
                elements.append(Paragraph("<b>LISTADO DETALLADO DE REQUERIMIENTOS</b>", subtitle_style))
                
                for req in requerimientos:
                    req_title = Paragraph(f"<b>REQ-{req.pk}: {req.nombre}</b>", 
                                         ParagraphStyle('ReqTitle', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#2c3e50'), spaceAfter=4))
                    elements.append(req_title)
                    
                    req_info = f"<b>Tipo:</b> {req.get_tipo_display() if hasattr(req, 'get_tipo_display') else 'N/A'} | "
                    req_info += f"<b>Estado:</b> {req.get_estado_display() if hasattr(req, 'get_estado_display') else 'N/A'} | "
                    req_info += f"<b>Prioridad:</b> {req.get_prioridad_display() if hasattr(req, 'get_prioridad_display') else 'N/A'}"
                    
                    elements.append(Paragraph(req_info, ParagraphStyle('ReqInfo', parent=styles['Normal'], fontSize=9, textColor=colors.grey)))
                    
                    if req.descripcion:
                        elements.append(Paragraph(f"<b>Descripción:</b> {req.descripcion}", normal_style))
                    
                    elements.append(Spacer(1, 0.5*cm))
                
                elements.append(Spacer(1, 0.5*cm))
        
        # === LISTADO DE CASOS DE USO (opcional) ===
        if incluir_casos_uso:
            from casos_de_uso.models import CasoDeUso
            
            casos = CasoDeUso.objects.filter(proyecto=proyecto).order_by('id')
            
            if casos.exists():
                elements.append(PageBreak())
                elements.append(Paragraph("<b>LISTADO DETALLADO DE CASOS DE USO</b>", subtitle_style))
                elements.append(Spacer(1, 0.3*cm))
                
                for caso in casos:
                    # Título del caso de uso
                    caso_title = Paragraph(
                        f"<b>CU-{caso.pk}: {caso.nombre}</b>", 
                        ParagraphStyle('CasoTitle', parent=styles['Normal'], fontSize=11, 
                                     textColor=colors.HexColor('#2980b9'), spaceAfter=6)
                    )
                    elements.append(caso_title)
                    
                    # Información básica
                    if caso.descripcion:
                        elements.append(Paragraph(f"<b>Descripción:</b> {caso.descripcion}", normal_style))
                        elements.append(Spacer(1, 0.2*cm))
                    
                    # Requerimientos relacionados
                    reqs_relacionados = caso.requerimientos_relacionados.all()
                    if reqs_relacionados.exists():
                        reqs_texto = ", ".join([f"REQ-{req.pk}" for req in reqs_relacionados])
                        elements.append(Paragraph(
                            f"<b>Requerimientos Relacionados:</b> {reqs_texto}", 
                            normal_style
                        ))
                        elements.append(Spacer(1, 0.2*cm))
                    
                    # Usuario creador y fechas
                    info_metadata = f"<b>Creado por:</b> {caso.creado_por.nombre if caso.creado_por else 'N/A'} | "
                    info_metadata += f"<b>Fecha creación:</b> {caso.fecha_creacion.strftime('%d/%m/%Y')}"
                    elements.append(Paragraph(
                        info_metadata, 
                        ParagraphStyle('Metadata', parent=styles['Normal'], fontSize=8, 
                                     textColor=colors.grey, spaceAfter=6)
                    ))
                    
                    # DETALLE TRADICIONAL
                    if hasattr(caso, 'detalle_tradicional') and caso.detalle_tradicional:
                        detalle = caso.detalle_tradicional
                        
                        elements.append(Paragraph(
                            "<b><i>Detalle Metodología Tradicional:</i></b>", 
                            ParagraphStyle('SubHeader', parent=styles['Normal'], fontSize=10, 
                                         textColor=colors.HexColor('#34495e'), spaceAfter=4)
                        ))
                        
                        if detalle.actor_principal:
                            elements.append(Paragraph(
                                f"<b>Actor Principal:</b> {detalle.actor_principal}", 
                                normal_style
                            ))
                            elements.append(Spacer(1, 0.1*cm))
                        
                        if detalle.precondiciones:
                            elements.append(Paragraph(
                                f"<b>Precondiciones:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.precondiciones, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                        
                        if detalle.flujo_principal:
                            elements.append(Paragraph(
                                f"<b>Flujo Principal:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.flujo_principal, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                        
                        if detalle.flujo_alternativo:
                            elements.append(Paragraph(
                                f"<b>Flujo Alternativo:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.flujo_alternativo, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                        
                        if detalle.postcondiciones:
                            elements.append(Paragraph(
                                f"<b>Postcondiciones:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.postcondiciones, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                        
                        if detalle.observaciones:
                            elements.append(Paragraph(
                                f"<b>Observaciones:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.observaciones, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                    
                    # DETALLE ÁGIL
                    elif hasattr(caso, 'detalle_agil') and caso.detalle_agil:
                        detalle = caso.detalle_agil
                        
                        elements.append(Paragraph(
                            "<b><i>Detalle Metodología Ágil:</i></b>", 
                            ParagraphStyle('SubHeader', parent=styles['Normal'], fontSize=10, 
                                         textColor=colors.HexColor('#27ae60'), spaceAfter=4)
                        ))
                        
                        if detalle.historia_usuario:
                            elements.append(Paragraph(
                                f"<b>Historia de Usuario:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.historia_usuario, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                        
                        if detalle.criterio_aceptacion:
                            elements.append(Paragraph(
                                f"<b>Criterios de Aceptación:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.criterio_aceptacion, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                        
                        if detalle.responsable:
                            elements.append(Paragraph(
                                f"<b>Responsable:</b> {detalle.responsable}", 
                                normal_style
                            ))
                            elements.append(Spacer(1, 0.1*cm))
                        
                        if detalle.estado_scrum:
                            elements.append(Paragraph(
                                f"<b>Estado Scrum:</b> {detalle.estado_scrum}", 
                                normal_style
                            ))
                            elements.append(Spacer(1, 0.1*cm))
                        
                        if detalle.observaciones:
                            elements.append(Paragraph(
                                f"<b>Observaciones:</b>", 
                                normal_style
                            ))
                            elements.append(Paragraph(
                                detalle.observaciones, 
                                ParagraphStyle('Detail', parent=styles['Normal'], fontSize=9, 
                                             leftIndent=20, spaceAfter=4)
                            ))
                    
                    else:
                        elements.append(Paragraph(
                            "<i>Este caso de uso no tiene detalle metodológico asignado.</i>", 
                            ParagraphStyle('Note', parent=styles['Normal'], fontSize=9, 
                                         textColor=colors.grey, italic=True, spaceAfter=4)
                        ))
                    
                    # Link externo si existe
                    if caso.link_externo:
                        elements.append(Paragraph(
                            f"<b>Recurso Externo:</b> <link href='{caso.link_externo}'>{caso.link_externo}</link>", 
                            normal_style
                        ))
                        elements.append(Spacer(1, 0.1*cm))
                    
                    # Separador entre casos de uso
                    elements.append(Spacer(1, 0.5*cm))
                    elements.append(Paragraph(
                        "─" * 100, 
                        ParagraphStyle('Separator', parent=styles['Normal'], fontSize=8, 
                                     textColor=colors.lightgrey)
                    ))
                    elements.append(Spacer(1, 0.5*cm))
                
                elements.append(Spacer(1, 0.3*cm))
        
        # === INFORMACIÓN DEL GRUPO (opcional) ===
        if incluir_info_grupo and proyecto.grupo:
            elements.append(Paragraph("<b>INFORMACIÓN DEL GRUPO</b>", subtitle_style))
            
            grupo_info = f"<b>Nombre:</b> {proyecto.grupo.nombre}<br/>"
            
            # Líder del grupo
            if proyecto.grupo.lider:
                grupo_info += f"<b>Líder del Grupo:</b> {proyecto.grupo.lider.nombre}<br/>"
            
            # Integrantes
            integrantes = proyecto.grupo.integrantes.all()
            grupo_info += f"<b>Cantidad de Integrantes:</b> {integrantes.count()}<br/>"
            
            # Estado
            grupo_info += f"<b>Estado:</b> {'Activo' if proyecto.grupo.activo else 'Inactivo'}"
            
            elements.append(Paragraph(grupo_info, normal_style))
            elements.append(Spacer(1, 1*cm))
        
        # === RECOMENDACIONES (opcional) ===
        if incluir_recomendaciones:
            from requerimientos.models import Requerimiento
            from casos_de_uso.models import CasoDeUso
            
            requerimientos = Requerimiento.objects.filter(proyecto=proyecto)
            casos = CasoDeUso.objects.filter(proyecto=proyecto)
            
            total_reqs = requerimientos.count()
            total_casos = casos.count()
            reqs_con_casos = sum(1 for req in requerimientos if req.casos_relacionados.exists())
            casos_con_reqs = sum(1 for caso in casos if caso.requerimientos_relacionados.exists())
            
            recomendaciones = []
            
            # === ANÁLISIS DE TRAZABILIDAD ===
            if total_reqs - reqs_con_casos > 0:
                recomendaciones.append(
                    f"• <b>Trazabilidad:</b> Hay {total_reqs - reqs_con_casos} "
                    f"requerimiento(s) sin casos de uso asociados. Se recomienda "
                    f"vincularlos para mejorar la trazabilidad."
                )
            
            if total_casos - casos_con_reqs > 0:
                recomendaciones.append(
                    f"• <b>Trazabilidad:</b> Hay {total_casos - casos_con_reqs} "
                    f"caso(s) de uso sin requerimientos asociados. Verificar si son "
                    f"necesarios o vincularlos."
                )
            
            if reqs_con_casos == total_reqs and total_reqs > 0:
                recomendaciones.append(
                    "• ✓ <b>Excelente:</b> Todos los requerimientos tienen casos "
                    "de uso asociados."
                )
            
            if casos_con_reqs == total_casos and total_casos > 0:
                recomendaciones.append(
                    "• ✓ <b>Excelente:</b> Todos los casos de uso están vinculados "
                    "a requerimientos."
                )
            
            # === ANÁLISIS DE ESTADOS ===
            borrador_count = requerimientos.filter(estado='BORRADOR').count()
            validado_count = requerimientos.filter(estado='VALIDADO').count()
            en_proceso_count = requerimientos.filter(estado='EN_PROCESO').count()
            terminado_count = requerimientos.filter(estado='TERMINADO').count()
            
            # Detectar muchos requerimientos en BORRADOR
            if total_reqs > 0 and borrador_count > total_reqs * 0.5:
                porcentaje_borrador = round((borrador_count / total_reqs) * 100, 1)
                recomendaciones.append(
                    f"⚠ <b>Advertencia:</b> {borrador_count} requerimientos "
                    f"({porcentaje_borrador}%) están en estado BORRADOR. "
                    f"Considere validarlos para avanzar en el proyecto."
                )
            
            # Detectar pocos requerimientos terminados
            if total_reqs > 0 and terminado_count < total_reqs * 0.2 and en_proceso_count > 0:
                porcentaje_terminado = round((terminado_count / total_reqs) * 100, 1)
                recomendaciones.append(
                    f"• <b>Progreso:</b> Solo {terminado_count} requerimientos "
                    f"({porcentaje_terminado}%) están terminados. Considere revisar "
                    f"los {en_proceso_count} en proceso para completarlos."
                )
            
            # Felicitar por buen progreso
            if total_reqs > 0 and terminado_count >= total_reqs * 0.8:
                porcentaje_terminado = round((terminado_count / total_reqs) * 100, 1)
                recomendaciones.append(
                    f"• ✓ <b>Excelente progreso:</b> {terminado_count} requerimientos "
                    f"({porcentaje_terminado}%) completados. ¡El proyecto avanza bien!"
                )
            
            # === ANÁLISIS DE BALANCE RF vs RNF ===
            rf_count = requerimientos.filter(tipo='FUNCIONAL').count()
            rnf_count = requerimientos.filter(tipo='NO_FUNCIONAL').count()
            rs_count = requerimientos.filter(tipo='SISTEMA').count()
            
            if rnf_count == 0 and rf_count > 5:
                recomendaciones.append(
                    "⚠ <b>Requerimientos No Funcionales:</b> No se han definido "
                    "RNF. Considere agregar requisitos de rendimiento, seguridad, "
                    "usabilidad, etc."
                )
            
            if rnf_count > 0 and rf_count > 0:
                ratio = rf_count / rnf_count if rnf_count > 0 else 0
                if ratio > 10:  # Más de 10 RF por cada RNF
                    recomendaciones.append(
                        f"• <b>Balance RF/RNF:</b> Hay {rf_count} RF y solo {rnf_count} RNF "
                        f"(ratio {ratio:.1f}:1). Considere revisar si faltan requisitos "
                        f"no funcionales importantes."
                    )
            
            # === ANÁLISIS DE PRIORIZACIÓN ===
            sin_prioridad_count = 0
            for req in requerimientos:
                if hasattr(req, 'detalle_tradicional') and req.detalle_tradicional:
                    if not req.detalle_tradicional.prioridad:
                        sin_prioridad_count += 1
                elif hasattr(req, 'detalle_agil') and req.detalle_agil:
                    if not req.detalle_agil.prioridad:
                        sin_prioridad_count += 1
            
            if sin_prioridad_count > 0:
                porcentaje_sin_prioridad = round((sin_prioridad_count / total_reqs) * 100, 1) if total_reqs > 0 else 0
                recomendaciones.append(
                    f"• <b>Priorización:</b> {sin_prioridad_count} requerimientos "
                    f"({porcentaje_sin_prioridad}%) no tienen prioridad asignada. "
                    f"Se recomienda priorizarlos usando MoSCoW (Must/Should/Could/Won't)."
                )
            
            # === ANÁLISIS DE FECHAS DE COMPROMISO ===
            from django.utils import timezone
            sin_fecha_count = 0
            fechas_vencidas = 0
            
            for req in requerimientos:
                if hasattr(req, 'detalle_tradicional') and req.detalle_tradicional:
                    if not req.detalle_tradicional.fecha_compromiso:
                        sin_fecha_count += 1
                    elif req.detalle_tradicional.fecha_compromiso < timezone.now().date() and req.estado != 'TERMINADO':
                        fechas_vencidas += 1
            
            if sin_fecha_count > total_reqs * 0.3 and total_reqs > 0:
                porcentaje_sin_fecha = round((sin_fecha_count / total_reqs) * 100, 1)
                recomendaciones.append(
                    f"• <b>Planificación:</b> {sin_fecha_count} requerimientos "
                    f"({porcentaje_sin_fecha}%) no tienen fecha de compromiso. "
                    f"Asignar fechas ayuda a gestionar expectativas."
                )
            
            if fechas_vencidas > 0:
                recomendaciones.append(
                    f"⚠ <b>Atención:</b> {fechas_vencidas} requerimiento(s) tienen "
                    f"fecha de compromiso vencida y no están terminados. "
                    f"Considere revisar el cronograma."
                )
            
            # === ANÁLISIS DE DEPENDENCIAS ===
            reqs_con_dependencias = sum(1 for req in requerimientos if req.dependencias.exists())
            if reqs_con_dependencias > 0:
                recomendaciones.append(
                    f"• <b>Dependencias:</b> {reqs_con_dependencias} requerimientos "
                    f"tienen dependencias. Asegúrese de implementarlos en el orden correcto."
                )
            
            # === RECOMENDACIÓN GENERAL SI NO HAY DATOS ===
            if total_reqs == 0:
                recomendaciones.append(
                    "• <b>Inicio del proyecto:</b> No hay requerimientos registrados. "
                    "Comience definiendo los requerimientos funcionales y no funcionales."
                )
            
            if total_casos == 0 and total_reqs > 0:
                recomendaciones.append(
                    "• <b>Casos de Uso:</b> No hay casos de uso definidos. Considere "
                    "crear casos de uso para detallar cómo los usuarios interactuarán "
                    "con el sistema."
                )
            
            if recomendaciones:
                elements.append(Paragraph("<b>RECOMENDACIONES Y ANÁLISIS</b>", subtitle_style))
                for rec in recomendaciones:
                    elements.append(Paragraph(rec, normal_style))
                elements.append(Spacer(1, 1*cm))
        
        # === CONSTRUIR PDF CON HEADER Y FOOTER ===
        # Usar onPage callback para agregar header/footer en cada página
        def add_page_decorations(canvas, doc):
            crear_header_footer(canvas, doc, proyecto, logo_proyecto_path, logo_grupo_path)
        
        doc.build(elements, onFirstPage=add_page_decorations, onLaterPages=add_page_decorations)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{proyecto.nombre}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        return response
        
    except ImportError as e:
        messages.error(request, f"Error de importación al generar el reporte: {str(e)}")
        return redirect('proyectos:proyecto_reportes', proyecto_id=proyecto_id)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR AL GENERAR REPORTE: {error_detail}")
        messages.error(request, f"Error al generar el reporte: {str(e)}")
        return redirect('proyectos:proyecto_reportes', proyecto_id=proyecto_id)


@login_required
def reportes_lider(request):
    """
    Redirige al líder a la página de reportes de su proyecto.
    Si tiene múltiples proyectos, redirige al primero.
    """
    proyectos = Proyecto.objects.filter(lider=request.user)
    
    if not proyectos.exists():
        messages.warning(request, "No tienes proyectos asignados como líder.")
        return redirect('dashboards:lider_dashboard')
    
    # Redirigir al reporte del primer proyecto
    proyecto = proyectos.first()
    return redirect('proyectos:proyecto_reportes', proyecto_id=proyecto.id)
